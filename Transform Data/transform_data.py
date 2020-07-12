import pandas as pd
import openpyxl as op
from datetime import datetime, timedelta

output_list_1 = [
    'RSTPSU1TO6',
    'RSTPSU7',
    'NLCIIST1',
    'NLCIIST2',
    'NLCEXP',
    'TALST2',
    'SIMHST2',
    'VALLURNTECL',
]
output_list_2 = [
    'NLCTS2EXP',
    'NTPL'
]
output_list_3 = [
    'SIMHST1',
    'KUDGI',
    'NNTPP'
]

input_files = ['curtailment_input_1.xlsx','curtailment_input_2.xlsx','curtailment_input_3.xlsx','curtailment_input_4.xlsx','curtailment_input_5.xlsx','curtailment_input_6.xlsx']

output_df = pd.DataFrame()

req_df_list = []
for req_file in input_files:
    req_file_xlsx = pd.ExcelFile(req_file)
    ############################ make file to check whether sheeet date and data date are correct ################
    wb = op.load_workbook(req_file)
    sheets = wb.sheetnames
    for sheet in sheets:
        req_df = pd.DataFrame()
        dt = datetime.strptime(sheet.split('_')[1],'%d-%m-%Y')
        df = pd.read_excel(req_file_xlsx, sheet_name = sheet,skiprows = 4,header = [0,1])
        if dt.strftime('%d-%m-%Y') not in wb[sheet]['A2'].value:
            print(dt.strftime('%d-%m-%Y')+"jkljlkjlkjkljkjkljl;kjkljkljkjkjjkjlk")
            continue
        del df['Time Desc']
        req_df['Date'] = [dt.strftime('%d-%m-%Y') for x in range(96)]
        req_df['Block'] = [x for x in range(1,97)]
        df = df.iloc[0:96].copy()
        for gen_obj in output_list_1:
            gen_df = df[gen_obj]
            declaredRUP = gen_df['RUP'].tolist()
            declaredRDOWN = gen_df['RDOWN'].tolist()
            req_df[gen_obj + '_RAMP'] = [min(declaredRUP[i],declaredRDOWN[i]) for i in range(96)]
            req_df[gen_obj + '_NDC'] = gen_df['NORMATIVE_DC'].tolist()
        req_df_list.append(req_df)
        print(dt.strftime('%d-%m-%Y'))
final_req_df = pd.DataFrame(columns = req_df.columns)
for df in req_df_list:
    final_req_df = final_req_df.append(df)

############################### for second list ################################################################3
req_df_list = []
for req_file in input_files:
    req_file_xlsx = pd.ExcelFile(req_file)
    ############################ make file to check whether sheeet date and data date are correct ################
    wb = op.load_workbook(req_file)
    sheets = wb.sheetnames
    for sheet in sheets:
        req_df = pd.DataFrame()
        dt = datetime.strptime(sheet.split('_')[1],'%d-%m-%Y')
        df = pd.read_excel(req_file_xlsx, sheet_name = sheet,skiprows = 4,header = [0,1])
        if dt.strftime('%d-%m-%Y') not in wb[sheet]['A2'].value:
            print(dt.strftime('%d-%m-%Y'))
            continue
        del df['Time Desc']
##        req_df['Date'] = [dt.strftime('%d-%m-%Y') for x in range(96)]
##        req_df['Block'] = [x for x in range(1,97)]
        df = df.iloc[0:96].copy()
        for gen_obj in output_list_2:
            gen_df = df[gen_obj]
            declaredRUP = gen_df['RUP'].tolist()
            declaredRDOWN = gen_df['RDOWN'].tolist()
            req_df[gen_obj + '_RAMP'] = [min(declaredRUP[i],declaredRDOWN[i]) for i in range(96)]
            req_df[gen_obj + '_NDC'] = gen_df['NORMATIVE_DC'].tolist()
        req_df_list.append(req_df)
        print(dt.strftime('%d-%m-%Y'))
final_req_df_2 = pd.DataFrame(columns = req_df.columns)
for df in req_df_list:
    final_req_df_2 = final_req_df_2.append(df)

############################### for third list ################################################################
req_df_list = []
for req_file in input_files:
    req_file_xlsx = pd.ExcelFile(req_file)
    ############################ make file to check whether sheeet date and data date are correct ################
    wb = op.load_workbook(req_file)
    sheets = wb.sheetnames
    for sheet in sheets:
        req_df = pd.DataFrame()
        dt = datetime.strptime(sheet.split('_')[1],'%d-%m-%Y')
        df = pd.read_excel(req_file_xlsx, sheet_name = sheet,skiprows = 4,header = [0,1])
        if dt.strftime('%d-%m-%Y') not in wb[sheet]['A2'].value:
            print(dt.strftime('%d-%m-%Y'))
            continue
        del df['Time Desc']
##        req_df['Date'] = [dt.strftime('%d-%m-%Y') for x in range(96)]
##        req_df['Block'] = [x for x in range(1,97)]
        df = df.iloc[0:96].copy()
        for gen_obj in output_list_3:
            gen_df = df[gen_obj]
            declaredRUP = gen_df['RUP'].tolist()
            declaredRDOWN = gen_df['RDOWN'].tolist()
            req_df[gen_obj + '_RAMP'] = [min(declaredRUP[i],declaredRDOWN[i]) for i in range(96)]
            req_df[gen_obj + '_NDC'] = gen_df['NORMATIVE_DC'].tolist()
        req_df_list.append(req_df)
        print(dt.strftime('%d-%m-%Y'))
final_req_df_3 = pd.DataFrame(columns = req_df.columns)
for df in req_df_list:
    final_req_df_3 = final_req_df_3.append(df)

template_file = op.load_workbook('template.xlsx')
writer = pd.ExcelWriter('result.xlsx', engine='openpyxl')
writer.book = template_file
writer.sheets = dict((ws.title, ws) for ws in template_file.worksheets)
final_req_df.to_excel(writer,sheet_name = 'RAMP',startrow=0,startcol=0,index=False)
final_req_df_2.to_excel(writer,sheet_name = 'RAMP',startrow=0,startcol=23,index=False)
final_req_df_3.to_excel(writer,sheet_name = 'RAMP',startrow=0,startcol=28,index=False)
writer.save()
import pdb
pdb.set_trace()

